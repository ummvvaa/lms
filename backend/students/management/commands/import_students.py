"""Импорт учеников из CSV/XLSX.

Без магии: явный маппинг колонок, отчёт об ошибках, идемпотентность по email.
Повторный запуск того же файла обновляет существующих учеников и не плодит
дубликатов. Каждое изменение доменного поля пишется в AuditLog с источником
`import` (инвариант №9).
"""

from __future__ import annotations

import csv
from collections.abc import Iterator
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.audit import apply_changes
from core.domains import Source
from students.models import (
    AdmissionProfile,
    BehaviorProfile,
    ExamProfile,
    SportProfile,
    Student,
    StudyGroup,
    TalentProfile,
)

#: Колонка файла → (модель-профиль, поле). Реестровые поля ученика — под ключом None.
COLUMN_MAP: dict[str, tuple[str | None, str]] = {
    "last_name": (None, "last_name"),
    "first_name": (None, "first_name"),
    "middle_name": (None, "middle_name"),
    "email": (None, "email"),
    "grade": (None, "grade"),
    "group": (None, "group"),
    "graduation_year": (None, "graduation_year"),
    "attendance_percent": ("behavior", "attendance_percent"),
    "remarks_count": ("behavior", "remarks_count"),
    "homework_percent": ("behavior", "homework_percent"),
    "target_country": ("admission", "target_country"),
    "target_major": ("admission", "target_major"),
    "has_common_app": ("admission", "has_common_app"),
    "has_application_account": ("admission", "has_application_account"),
    "ielts_current": ("exam", "ielts_current"),
    "ielts_target": ("exam", "ielts_target"),
    "sat_current": ("exam", "sat_current"),
    "sat_target": ("exam", "sat_target"),
    "hours_per_week": ("exam", "hours_per_week"),
    "teacher": ("exam", "teacher"),
    "gpa": ("exam", "gpa"),
    "main_track": ("talent", "main_track"),
    "sport_kind": ("sport", "sport_kind"),
    "level": ("sport", "level"),
    "rank": ("sport", "rank"),
}

PROFILE_MODELS = {
    "behavior": BehaviorProfile,
    "admission": AdmissionProfile,
    "exam": ExamProfile,
    "talent": TalentProfile,
    "sport": SportProfile,
}

REQUIRED_COLUMNS = ("last_name", "first_name", "email", "grade", "graduation_year")

INT_FIELDS = {
    "grade",
    "graduation_year",
    "attendance_percent",
    "remarks_count",
    "homework_percent",
    "sat_current",
    "sat_target",
    "hours_per_week",
}
DECIMAL_FIELDS = {"ielts_current", "ielts_target", "gpa"}
BOOL_FIELDS = {"has_common_app", "has_application_account"}

TRUE_WORDS = {"1", "true", "yes", "да", "y", "+", "есть"}
FALSE_WORDS = {"0", "false", "no", "нет", "n", "-", ""}


@dataclass
class ImportReport:
    """Что получилось: сколько создано, сколько обновлено, где ошибки."""

    created: int = 0
    updated: int = 0
    unchanged: int = 0
    audit_entries: int = 0
    errors: list[str] = field(default_factory=list)

    def as_text(self) -> str:
        lines = [
            f"Создано учеников: {self.created}",
            f"Обновлено: {self.updated}",
            f"Без изменений: {self.unchanged}",
            f"Записей в аудит: {self.audit_entries}",
            f"Ошибок: {len(self.errors)}",
        ]
        lines += [f"  строка {e}" for e in self.errors]
        return "\n".join(lines)


def read_rows(path: Path) -> Iterator[dict[str, str]]:
    """Прочитать CSV или XLSX как последовательность словарей."""
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        for row in rows:
            yield {header[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(row) if i < len(header)}
        wb.close()
    else:
        with path.open(encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                yield {(k or "").strip(): (v or "").strip() for k, v in row.items()}


def coerce(field_name: str, raw: str) -> Any:
    """Привести значение колонки к типу поля модели."""
    if field_name in BOOL_FIELDS:
        low = raw.strip().lower()
        if low in TRUE_WORDS:
            return True
        if low in FALSE_WORDS:
            return False
        raise ValueError(f"{field_name}: не понял «{raw}» как да/нет")
    if raw == "":
        return None
    if field_name in INT_FIELDS:
        try:
            return int(float(raw.replace(",", ".")))
        except ValueError as exc:
            raise ValueError(f"{field_name}: ожидалось число, получено «{raw}»") from exc
    if field_name in DECIMAL_FIELDS:
        try:
            return Decimal(raw.replace(",", "."))
        except InvalidOperation as exc:
            raise ValueError(f"{field_name}: ожидалось число, получено «{raw}»") from exc
    return raw


class Command(BaseCommand):
    help = "Импортирует учеников и профили из CSV/XLSX. Идемпотентно по email."

    def add_arguments(self, parser) -> None:
        parser.add_argument("path", type=str, help="Путь к CSV или XLSX")
        parser.add_argument("--dry-run", action="store_true", help="Только проверить файл, ничего не писать")

    def handle(self, *args, **options) -> None:
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"Файл не найден: {path}")
        report = run_import(path, dry_run=options["dry_run"])
        self.stdout.write(report.as_text())
        if report.errors:
            self.stdout.write(self.style.WARNING("Импорт завершён с ошибками — см. список выше"))
        else:
            self.stdout.write(self.style.SUCCESS("Импорт завершён"))


def run_import(path: Path, *, dry_run: bool = False, actor=None) -> ImportReport:
    """Основная работа импорта, вынесена из команды ради тестов и API."""
    report = ImportReport()
    rows = list(read_rows(path))
    if not rows:
        report.errors.append("0: файл пустой")
        return report

    missing = [c for c in REQUIRED_COLUMNS if c not in rows[0]]
    if missing:
        report.errors.append(f"1: нет обязательных колонок: {', '.join(missing)}")
        return report

    with transaction.atomic():
        for number, row in enumerate(rows, start=2):
            try:
                _import_row(row, report, actor=actor)
            except Exception as exc:
                report.errors.append(f"{number}: {exc}")
        if dry_run:
            transaction.set_rollback(True)
    return report


def _import_row(row: dict[str, str], report: ImportReport, *, actor=None) -> None:
    """Импорт одной строки: ученик, затем его профили."""
    email = (row.get("email") or "").strip().lower()
    if not email:
        raise ValueError("пустой email — строку пропускаем")

    student_values: dict[str, Any] = {}
    profile_values: dict[str, dict[str, Any]] = {name: {} for name in PROFILE_MODELS}

    for column, raw in row.items():
        target = COLUMN_MAP.get(column)
        if target is None:
            continue
        profile_name, field_name = target
        if field_name == "email":
            continue
        if field_name == "group":
            student_values["group"] = _group_for(raw, row) if raw else None
            continue
        value = coerce(field_name, raw)
        if value is None and field_name in ("last_name", "first_name", "grade", "graduation_year"):
            raise ValueError(f"{field_name}: обязательное значение пустое")
        if profile_name is None:
            if value is not None:
                student_values[field_name] = value
        elif value is not None:
            profile_values[profile_name][field_name] = value

    student = Student.objects.filter(email=email).first()
    created = student is None
    changed = 0
    if created:
        student = Student(email=email, **student_values)
        student.save()
        report.created += 1
    else:
        entries = apply_changes(student, student_values, actor=actor, source=Source.IMPORT)
        report.audit_entries += len(entries)
        changed += len(entries)

    for name, model in PROFILE_MODELS.items():
        values = profile_values[name]
        profile, _ = model.objects.get_or_create(student=student)
        if values:
            entries = apply_changes(profile, values, actor=actor, source=Source.IMPORT)
            report.audit_entries += len(entries)
            changed += len(entries)

    if not created:
        if changed:
            report.updated += 1
        else:
            report.unchanged += 1


def _group_for(code: str, row: dict[str, str]) -> StudyGroup:
    """Группа по коду; если её нет — заводим с классом из строки."""
    grade = int(float(row.get("grade") or 0)) or 10
    group, _ = StudyGroup.objects.get_or_create(code=code.strip(), defaults={"grade": grade})
    return group
