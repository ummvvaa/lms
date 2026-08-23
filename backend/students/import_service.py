"""Импорт из файла через интерфейс: предпросмотр, сопоставление, применение.

Отличается от management-команды тем, что колонки директор сопоставляет
руками, а править ему разрешено только свой домен (инвариант №1).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

from django.apps import apps
from django.db import transaction

from core.audit import ValueRejected, apply_changes, coerce, normalize, to_text
from core.domains import Source, can_write, domain_of_role, spec_of_field
from students.models import Student

MAX_PREVIEW_ROWS = 20


def read_table(uploaded) -> tuple[list[str], list[list[str]]]:
    """Прочитать загруженный CSV/XLSX в заголовок и строки."""
    name = (uploaded.name or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(uploaded, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows, [])]
        body = [["" if v is None else str(v).strip() for v in row] for row in rows]
        wb.close()
        return header, body

    import csv
    import io

    text = uploaded.read()
    if isinstance(text, bytes):
        text = text.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return [c.strip() for c in rows[0]], [[c.strip() for c in r] for r in rows[1:]]


@dataclass
class ImportPreview:
    """Что система собирается сделать — до того, как это сделает.

    `problems` — разбор по строкам и колонкам: где, что не так и как
    исправить. Одна кривая строка не отменяет файл: остальные применяются,
    а её показываем отдельно, чтобы было что править.
    """

    columns: list[str]
    total_rows: int
    matched: int = 0
    unmatched: list[dict[str, Any]] = field(default_factory=list)
    conflicts: list[dict[str, Any]] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    problems: list[dict[str, Any]] = field(default_factory=list)

    @property
    def ready_rows(self) -> list[dict[str, Any]]:
        """Строки, которые можно применять: без ошибок в значениях."""
        broken = {p["row"] for p in self.problems}
        return [row for row in self.rows if row["row"] not in broken and row["changes"]]

    def as_dict(self) -> dict:
        return {
            "columns": self.columns,
            "total_rows": self.total_rows,
            "matched": self.matched,
            "unmatched": self.unmatched,
            "conflicts": self.conflicts,
            # предпросмотр показывает первые строки, применяются все
            "rows": self.rows[:MAX_PREVIEW_ROWS],
            "all_rows": self.rows,
            "errors": self.errors,
            "problems": self.problems,
            "ready": len(self.ready_rows),
            "broken": len({p["row"] for p in self.problems}),
        }


def _resolve_student(value: str) -> Student | None:
    """Ученик по email — на этой фазе только точное совпадение.

    Полноценное сопоставление по ФИО с оценкой уверенности — Фаза 5,
    там для этого отдельный сервис и разрешение неоднозначностей.
    """
    value = (value or "").strip().lower()
    if not value:
        return None
    return Student.objects.filter(email=value).first()


def build_preview(*, header: list[str], rows: list[list[str]], mapping: dict[str, str], role: str) -> ImportPreview:
    """Собрать предпросмотр: кого нашли, что поменяется, где конфликты.

    `mapping` — `{колонка файла: "app.Model.field"}`, плюс служебная
    колонка со значением `student` — по ней ищется ученик.
    """
    preview = ImportPreview(columns=header, total_rows=len(rows))
    index = {name: i for i, name in enumerate(header)}
    domain = domain_of_role(role)
    if domain is None:
        preview.errors.append("У вашей роли нет домена — импортировать нечего")
        return preview

    key_column = next((col for col, target in mapping.items() if target == "student"), None)
    if key_column is None:
        preview.errors.append(
            "Не указано, в какой колонке искать ученика. Выберите «Ученик (email)» "
            "у колонки с почтой — по ней строка находит своего человека"
        )
        return preview

    def cell_of(row: list[str], column: str) -> str:
        """Значение колонки в строке; отсутствующие колонки — пустая строка."""
        i = index.get(column)
        return row[i] if i is not None and i < len(row) else ""

    for number, row in enumerate(rows, start=2):

        def cell(column: str, _row: list[str] = row) -> str:
            return cell_of(_row, column)

        student = _resolve_student(cell(key_column))
        if student is None:
            preview.unmatched.append({"row": number, "value": cell(key_column)})
            continue

        changes: list[dict[str, Any]] = []
        for column, target in mapping.items():
            if target in ("student", "", None):
                continue
            try:
                app_label, model_name, field_name = target.rsplit(".", 2)
            except ValueError:
                preview.errors.append(
                    f"Колонка «{column}» сопоставлена с чем-то непонятным. Выберите поле из списка заново"
                )
                continue
            model_label = f"{app_label}.{model_name}"

            if not can_write(role, model_label, field_name):
                # чужой домен отсекается на сервере, а не прячется в интерфейсе
                preview.errors.append(
                    f"Колонка «{column}»: это поле ведёт другой директор. "
                    "Выберите для неё поле своего домена или не импортируйте её"
                )
                continue

            instance = apps.get_model(model_label).objects.filter(student=student).first()
            if instance is None:
                continue
            raw = cell(column)
            if raw == "":
                continue

            # значение проверяем здесь, до применения: строка 12 с IELTS 12.5
            # должна быть названа по имени, а не молча пропасть при сохранении
            try:
                coerce(instance, field_name, raw)
            except ValueRejected as problem:
                spec = spec_of_field(model_label, field_name)
                preview.problems.append(
                    {
                        "row": number,
                        "column": column,
                        "field": field_name,
                        "student_name": student.full_name,
                        "value": raw,
                        "message": str(problem),
                        "hint": spec.range_hint if spec else "",
                    }
                )
                continue

            old = to_text(getattr(instance, field_name, None))
            new = to_text(normalize(instance, field_name, raw))
            if old != new:
                changes.append({"model": model_label, "field": field_name, "old": old, "new": new, "raw": raw})
                if old:
                    preview.conflicts.append(
                        {"row": number, "student": student.pk, "field": field_name, "old": old, "new": new}
                    )

        preview.matched += 1
        preview.rows.append(
            {"row": number, "student": student.pk, "student_name": student.full_name, "changes": changes}
        )

    return preview


@transaction.atomic
def apply_preview(
    *,
    preview_rows: list[dict[str, Any]],
    role: str,
    actor=None,
    file_name: str = "",
) -> dict:
    """Применить то, что директор увидел в предпросмотре.

    Каждая загрузка заводит `ImportBatch`, и все её изменения ссылаются
    на него — по этой ссылке загрузку потом можно отменить целиком.
    """
    from core.models import ImportBatch

    domain = domain_of_role(role)
    batch = ImportBatch.objects.create(
        actor=actor,
        file_name=file_name,
        kind=ImportBatch.Kind.STUDENTS,
        domain_code=domain.code if domain else "",
        rows_total=len(preview_rows),
    )

    applied = 0
    audit_entries = 0
    touched_rows = 0
    rejected: list[dict[str, Any]] = []

    for row in preview_rows:
        student_id = row.get("student")
        grouped: dict[str, dict[str, Any]] = {}
        for change in row.get("changes", []):
            model_label, field_name = change["model"], change["field"]
            if not can_write(role, model_label, field_name):
                continue
            grouped.setdefault(model_label, {})[field_name] = change.get("raw", change.get("new"))

        for model_label, values in grouped.items():
            instance = apps.get_model(model_label).objects.filter(student_id=student_id).first()
            if instance is None:
                continue
            # мусор в клетке файла отклоняет строку, а не роняет весь импорт
            clean: dict[str, Any] = {}
            for field_name, raw in values.items():
                try:
                    clean[field_name] = coerce(instance, field_name, raw)
                except ValueRejected as error:
                    rejected.append({"student": student_id, "field": field_name, "reason": str(error)})
            if not clean:
                continue
            entries = apply_changes(instance, clean, actor=actor, source=Source.IMPORT, import_batch=batch)
            audit_entries += len(entries)
            applied += len(clean)
            if entries:
                touched_rows += 1

    batch.rows_updated = touched_rows
    batch.rows_failed = len({row["student"] for row in rejected})
    batch.save(update_fields=["rows_updated", "rows_failed"])

    return {
        "applied": applied,
        "audit_entries": audit_entries,
        "rejected": rejected,
        "batch": batch.pk,
        "detail": (
            f"Загрузка сохранена: изменено полей {applied} у {touched_rows} учеников. "
            "Отменить её целиком можно в истории загрузок"
        ),
    }
